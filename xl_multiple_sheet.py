import openpyxl
workbook = openpyxl.Workbook()

ws1 = 'ONE'
ws2 = 'TWO'

workbook.create_sheet(ws1)
workbook.create_sheet(ws2)

workbook[ws1]['A1'] = 'ID'
workbook[ws1]['B1'] = 'Text'
workbook[ws1]['C1'] = 'Category'
workbook[ws1]['A2'] = '234'
workbook[ws1]['B2'] = 'Sample'
workbook[ws1]['C2'] = 'ASD'

workbook[ws2]['A1'] = 'ID'
workbook[ws2]['B1'] = 'Text'
workbook[ws2]['C1'] = 'Category'
workbook[ws2]['A2'] = '566'
workbook[ws2]['B2'] = 'Sample'
workbook[ws2]['C2'] = 'PED'
workbook[ws2]['A3'] = '896'
workbook[ws2]['B3'] = 'SAmple'
workbook[ws2]['C3'] = 'HEAD'

# You can also call the line below to delete the sheet 'Sheet'
workbook.remove(workbook['ONE']) #Delete content from specific sheet
workbook.remove(workbook['Sheet']) #Remove the sheet from workbook

workbook.save('sam.xlsx')